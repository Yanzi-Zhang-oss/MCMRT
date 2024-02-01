#!/usr/bin/python3
 
from sklearn import * 
from openpyxl import *
from tkinter import * 
import tkinter.messagebox

def fetch_data_from_excel(path):

    try:
        wb = load_workbook(path)
    except:
        tkinter.messagebox.showinfo('messagebox','file is not supported!')
        return
    else:
        pass
        
    sheetlist = wb.sheetnames

    if len(sheetlist) == 1:
        sheet = wb[sheetlist[0]]
        data = sheet.cell(row=7, column=9) 
        print(data.value)
        pass
    

    